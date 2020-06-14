import openpyxl as xl
from openpyxl.utils.cell import range_boundaries
import re
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import points_to_pixels


class MyPyxl:
    def __init__(self, format_xlsx, pt_height=18.75, pix_width=72):
        self.wb = xl.load_workbook(format_xlsx)
        self.unit_pt_height = pt_height
        self.unit_pix_width = pix_width
        self.Range_xl = {}
        self.Defines_xl = {}
        for defines in self.wb.defined_names.definedName:
            if 'RNG' in defines.name:
                self.Range_xl[defines.name] = []
            elif 'CELL' in defines.name:
                self.Defines_xl[defines.name] = []
            elif 'COL' in defines.name:
                self.Defines_xl[defines.name] = []
            elif 'ROW' in defines.name:
                self.Defines_xl[defines.name] = []
            elif 'MTR' in defines.name:
                self.Defines_xl[defines.name] = []
            elif 'IMG' in defines.name:
                self.Defines_xl[defines.name] = []
    
    def get_Defines(self):
        return {key:[] for key in self.Defines_xl.keys()}
    
    def regist_Defines2Range(self, label,defs):
        self.Range_xl[label].append(defs)

    def write_cell(self, ws, ColRow, value):
        """Write the value to a cell
        
        Parameters
        -------
        ws : openpyxl.worksheet.worksheet.Worksheet
            worksheet
        ColRow : list[int]
            range of the cell
        value : str or int
            Value to be written to the cell

        """
        for row in ws.iter_rows(min_col=ColRow[0],
                                min_row=ColRow[1],
                                max_col=ColRow[2],
                                max_row=ColRow[3]):
            for cell in row:
                cell.value = value


    def write_list_col(self, ws, ColRow, list_values):
        """Write the value to cells in the vertical columns
        
        Parameters
        -------
        ws : openpyxl.worksheet.worksheet.Worksheet
            worksheet
        ColRow : list[int]
            range of the cell
        list_values : list
            Values to be written to cells

        """
        for row, value in zip(ws.iter_rows(min_col=ColRow[0],
                                        min_row=ColRow[1],
                                        max_col=ColRow[2],
                                        max_row=ColRow[3]),
                                        list_values):
            for cell in row:
                cell.value = value


    def write_list_row(self, ws, ColRow, list_values):
        """Write the value to cells in the horizontal rows
        
        Parameters
        -------
        ws : openpyxl.worksheet.worksheet.Worksheet
            worksheet
        ColRow : list[int]
            range of the cell
        list_values : list
            Values to be written to cells

        """
        for row in ws.iter_rows(min_col=ColRow[0],
                                min_row=ColRow[1],
                                max_col=ColRow[2],
                                max_row=ColRow[3]):
            for cell, value in zip(row, list_values):
                cell.value = value


    def write_list_matrix(self, ws, ColRow, mtr_values):
        """Write the value to cells in the horizontal rows
           and the vertical columns
        
        Parameters
        -------
        ws : openpyxl.worksheet.worksheet.Worksheet
            worksheet
        ColRow : list[int]
            range of the cell
        mtr_values : list
            Values to be written to cells

        """
        for row, list_col in zip(ws.iter_rows(min_col=ColRow[0],
                                        min_row=ColRow[1],
                                        max_col=ColRow[2],
                                        max_row=ColRow[3]),
                                        mtr_values):
            for cell, value in zip(row, list_col):
                cell.value = value


    def add_img(self, ws, ColRow, path_img):
        """Add an image
           The image size is automatically adjusted.
        
        Parameters
        -------
        ws : openpyxl.worksheet.worksheet.Worksheet
            worksheet
        ColRow : list[int]
            range of the cell
        path_img : str
            the path of the image data to be added

        """

        #Calculating the height of cells in a range with 
        #the number of cells and the default cell height
        height_pt = (ColRow[3] - ColRow[1] + 1) * self.unit_pt_height
        height_pix = points_to_pixels(height_pt)
        #Calculating the width of cells in a range with 
        #the number of cells and the default cell width
        width_pix = (ColRow[2] - ColRow[0] + 1) * self.unit_pix_width

        tmp_img = Image(path_img)

        aspct = tmp_img.height / tmp_img.width
        
        if aspct >= 1:
            tmp_img.height = height_pix
            tmp_img.width = height_pix / aspct
            if tmp_img.width > width_pix:
                tmp_img.width = width_pix
                tmp_img.height = tmp_img.width * aspct
        else:
            tmp_img.width = width_pix
            tmp_img.height = width_pix * aspct
            if tmp_img.height > height_pix:
                tmp_img.height = height_pix
                tmp_img.width = tmp_img.height / aspct

        anc = ws.cell(column=ColRow[0], row=ColRow[1]).coordinate
        ws.add_image(tmp_img, anc)


    def offset_colrow(self, ColRow, Right, below):
        """Offset the position of the cell
        
        Parameters
        -------
        ColRow : list[int]
            range of the cell
        Right : int
            value to offset Right
        below : int
            value to offset below
        

        Returns
        -------
        return new_colrow : list
            range of the cell
        """

        new_colrow = ColRow.copy()
        if Right > 0:
            len_col = ColRow[2] - ColRow[0]
            new_colrow[0] = ColRow[0] + len_col + Right
            new_colrow[2] = ColRow[2] + len_col + Right
        
        if below > 0:
            len_row = ColRow[3] - ColRow[1]
            new_colrow[1] = ColRow[1] + len_row + below
            new_colrow[3] = ColRow[3] + len_row + below

        return new_colrow

    def add_colrow_next_range(self, base_ColRow, ColRow, Right, below):
        """Offset the position of the cell to the next range
        
        Parameters
        -------
        base_ColRow : list[int]
            Position of the base range
        ColRow : list[int]
            range of the cell
        Right : int
            value to offset Right
        below : int
            value to offset below
        

        Returns
        -------
        return new_colrow : list
            range of the cell
        """
        new_colrow = ColRow.copy()

        if Right > 0:
            len_row = base_ColRow[2] - base_ColRow[0]
            new_colrow[0] = ColRow[0] + len_row
            new_colrow[2] = ColRow[2] + len_row 

        if below > 0:
            len_row = base_ColRow[3] - base_ColRow[1]
            new_colrow[1] = ColRow[1] + len_row
            new_colrow[3] = ColRow[3] + len_row

        return new_colrow


    def write_xl(self, dict_defines, base_ColRow, ofst_range_r, ofst_range_d):
        """Write data to Excel
        
        Parameters
        -------
        dict_defines : dict[str,any]
            Excel's Defined Names
        base_ColRow : list[int]
            Position of the base range
        ofst_range_r : int
            value to offset Right
        ofst_range_d : int
            value to offset below

        """
        for key in dict_defines.keys():
            offset_r = 0
            offset_d = 0

            #get the positon of Excel's Defined Names
            my_range = self.wb.defined_names[key]
            dests = my_range.destinations
            for title, coord in dests:
                ws = self.wb[title]
                ColRow = np.array(list(range_boundaries(coord)))
                ColRow = self.add_colrow_next_range(base_ColRow, ColRow, ofst_range_r, ofst_range_d)
            
            #if Defined Name contains 'LOOP', Repeat by shifting the number of range
            #DN means shift N below. RN means shift N right
            if 'LOOPD' in key:
                offset_d = int(re.match('LOOPD[0-9]*', key).group().replace('LOOPD', ''))
            if 'LOOPR' in key:
                offset_r = int(re.match('LOOPR[0-9]*', key).group().replace('LOOPR',''))

            loop_ColRow = ColRow.copy()

            #change the processing depending on Defined Name 
            for loop in dict_defines[key]:
                if 'CELL' in key:
                    self.write_cell(ws, loop_ColRow, loop)
                elif 'COL' in key:
                    self.write_list_col(ws, loop_ColRow, loop)
                elif 'ROW' in key:
                    self.write_list_row(ws, loop_ColRow, loop)
                elif 'MTR' in key:
                    self.write_list_matrix(ws, loop_ColRow, loop)
                elif 'IMG' in key:
                    self.add_img(ws, loop_ColRow, loop)
                loop_ColRow = self.offset_colrow(loop_ColRow, offset_r, offset_d)


    def create_xlsx(self, list_rng ,path_xlsx):
        for label in list_rng:
            base_rng = self.wb.defined_names[label]
            dests = base_rng.destinations
            offset_d = 0
            offset_r = 0
            for title, coord in dests:
                base_ColRow = np.array(list(range_boundaries(coord)))

            for defines in self.Range_xl[label]:
                self.write_xl(defines, base_ColRow, offset_r, offset_d)

                if 'RNGD' in label:
                    offset_d = int(re.match('RNGD[0-9]*', label).group().replace('RNGD', ''))
                else:
                    offset_d = 0
                if 'RNGR' in label:
                    offset_r = int(re.match('RNGR[0-9]*', label).group().replace('RNGR',''))
                else:
                    offset_r = 0

                base_ColRow = self.offset_colrow(base_ColRow, offset_r, offset_d)
                        

        self.wb.save(path_xlsx)
        self.wb.close()


if __name__ == '__main__':
    pass