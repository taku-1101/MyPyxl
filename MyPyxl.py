import openpyxl as xl
from openpyxl.utils.cell import range_boundaries
import numpy as np
import re
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import points_to_pixels


class MyPyxl:
    def __init__(self, format_xlsx, pt_height=18.75, pix_width=72):
        self.wb = xl.load_workbook(format_xlsx)
        self.unit_pt_height = pt_height
        self.unit_pix_width = pix_width
        self.Range_xl = {}
        self.Defines_xl = {}
        for defines in self.wb.defined_names.definedName:
            if 'RNG' in defines.name:
                self.Range_xl[defines.name] = []
            elif 'CELL' in defines.name:
                self.Defines_xl[defines.name] = []
            elif 'COL' in defines.name:
                self.Defines_xl[defines.name] = []
            elif 'ROW' in defines.name:
                self.Defines_xl[defines.name] = []
            elif 'MTR' in defines.name:
                self.Defines_xl[defines.name] = []
            elif 'IMG' in defines.name:
                self.Defines_xl[defines.name] = []
    
    def get_Defines(self):
        return dict.fromkeys(self.Defines_xl, [])
    
    def regist_Defines2Range(self, label,defs):
        self.Range_xl[label].append(defs)

    def write_cell(self, ws, ColRow, value):
        """1つのセルに値を書き込む
        """
        for row in ws.iter_rows(min_col=ColRow[0],
                                min_row=ColRow[1],
                                max_col=ColRow[2],
                                max_row=ColRow[3]):
            for cell in row:
                cell.value = value


    def write_list_col(self, ws, ColRow, list_txt):
        """複数のセルに値を書き込む
            列用
        """
        for row, value in zip(ws.iter_rows(min_col=ColRow[0],
                                        min_row=ColRow[1],
                                        max_col=ColRow[2],
                                        max_row=ColRow[3]),
                                        list_txt):
            for cell in row:
                cell.value = value


    def write_list_row(self, ws, ColRow, list_txt):
        """複数のセルに値を書き込む
            行用
        """
        for row in ws.iter_rows(min_col=ColRow[0],
                                min_row=ColRow[1],
                                max_col=ColRow[2],
                                max_row=ColRow[3]):
            for cell, value in zip(row, list_txt):
                cell.value = value


    def write_list_matrix(self, ws, ColRow, mtr_txt):
        """複数のセルに値を書き込む
            行列用
        """
        for row, list_txt in zip(ws.iter_rows(min_col=ColRow[0],
                                        min_row=ColRow[1],
                                        max_col=ColRow[2],
                                        max_row=ColRow[3]),
                                        mtr_txt):
            for cell, value in zip(row, list_txt):
                cell.value = value


    def write_img(self, ws, ColRow, img):
        """画像を指定のセルに合わせて挿入
        """

        #セル数とデフォルトのセルの高さ（point）で範囲内のセルの高さを求める
        height_pt = (ColRow[3] - ColRow[1] + 1) * self.unit_pt_height
        height_pix = points_to_pixels(height_pt)
        #セル数とデフォルトのセルの幅（pix）で範囲内のセルの高さを求める
        width_pix = (ColRow[2] - ColRow[0] + 1) * self.unit_pix_width

        add_img = Image(img)

        #縦横比を一定にしてリサイズ
        aspct = add_img.height / add_img.width
        #縦が長いとき
        if aspct >= 1:
            add_img.height = height_pix
            add_img.width = height_pix / aspct
        #横が長いとき
        else:
            add_img.width = width_pix
            add_img.height = width_pix * aspct

        anc = ws.cell(column=ColRow[0], row=ColRow[1]).coordinate
        ws.add_image(add_img, anc)


    def offset_cell(self, colrow, Right, down):
        new_colrow = colrow.copy()
        if Right > 0:
            len_col = colrow[2] - colrow[0]
            new_colrow[0] = colrow[0] + len_col + Right
            new_colrow[2] = colrow[2] + len_col + Right
        
        if down > 0:
            len_row = colrow[3] - colrow[1]
            new_colrow[1] = colrow[1] + len_row + down
            new_colrow[3] = colrow[3] + len_row + down

        return new_colrow

    def add_range(self, base, colrow, Right, down):
        new_colrow = colrow.copy()

        if Right > 0:
            len_row = base[2] - base[0]
            new_colrow[0] = colrow[0] + len_row
            new_colrow[2] = colrow[2] + len_row 

        if down > 0:
            len_row = base[3] - base[1]
            new_colrow[1] = colrow[1] + len_row
            new_colrow[3] = colrow[3] + len_row

        return new_colrow

    def write_xl(self, dict_defines, base_ColRow, ofst_range_r, ofst_range_d):
        for key in dict_defines.keys():
            offset_r = 0
            offset_d = 0
            my_range = self.wb.defined_names[key]
            dests = my_range.destinations
            for title, coord in dests:
                ws = self.wb[title]
                ColRow = np.array(list(range_boundaries(coord)))
                ColRow = self.add_range(base_ColRow, ColRow, ofst_range_r, ofst_range_d)
                
            if 'LOOPD' in key:
                offset_d = int(re.match('LOOPD[0-9]*', key).group().replace('LOOPD', ''))
            if 'LOOPR' in key:
                offset_r = int(re.match('LOOPR[0-9]*', key).group().replace('LOOPR',''))

            loop_ColRow = ColRow.copy()
            for loop in dict_defines[key]:
                if 'CELL' in key:
                    self.write_cell(ws, loop_ColRow, loop)
                elif 'COL' in key:
                    self.write_list_col(ws, loop_ColRow, loop)
                elif 'ROW' in key:
                    self.write_list_row(ws, loop_ColRow, loop)
                elif 'MTR' in key:
                    self.write_list_matrix(ws, loop_ColRow, loop)
                elif 'IMG' in key:
                    self.write_img(ws, loop_ColRow, loop)
                loop_ColRow = offset_cell(loop_ColRow, offset_r, offset_d)


    def create_xlsx(self, list_range ,path_xlsx):
        for label in list_range:
            base_range = self.wb.defined_names[label]
            dests = base_range.destinations
            offset_d = 0
            offset_r = 0
            for title, coord in dests:
                base_ColRow = np.array(list(range_boundaries(coord)))

            for defines in self.Range_xl[label]:
                self.write_xl(defines, base_ColRow, offset_r, offset_d)

                if 'RNGD' in label:
                    offset_d = int(re.match('RNGD[0-9]*', label).group().replace('RNGD', ''))
                else:
                    offset_d = 0
                if 'RNGR' in label:
                    offset_r = int(re.match('RNGR[0-9]*', label).group().replace('RNGR',''))
                else:
                    offset_r = 0

                base_ColRow = self.offset_cell(base_ColRow, offset_r, offset_d)
                        

        self.wb.save(path_xlsx)
        self.wb.close()
